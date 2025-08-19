#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import time
import platform
import socket
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (TimeoutException, StaleElementReferenceException,
                                        WebDriverException, NoSuchElementException)
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 初始化colorama，支持跨平台彩色输出
init(autoreset=True)


def print_banner():
    """打印优化后的banner图案，修复反斜杠转义警告"""
    # 清除屏幕以获得更好的展示效果
    os.system('cls' if os.name == 'nt' else 'clear')

    # 使用原始字符串(r前缀)避免反斜杠被解析为转义字符，解决SyntaxWarning
    banner_pattern = [
        r"  ____  _                  _     ___  ____        _  _  _ ",
        r" / ___|| |__    ___   ___ | | __|_ _||  _ \  ___ | || || |",
        r"| |    | '_ \  / _ \ / __|| |/ / | | | |_) |/ __|| || || |",
        r"| |___ | | | ||  __/| (__ |   <  | | |  __/ \__ \|_||_||_|",
        r"\____||_| |_| \___| \___||_|\_\|___||_|    |___/(_)(_)(_)"
    ]

    # 计算最长行的长度，用于居中对齐
    max_length = max(len(line) for line in banner_pattern)

    # 打印顶部边框
    print(Style.BRIGHT + Fore.YELLOW + "=" * (max_length + 4))
    print()

    # 打印banner图案，确保居中对齐
    for line in banner_pattern:
        # 计算每行的缩进，实现居中显示
        indent = (max_length - len(line)) // 2
        print(Style.BRIGHT + Fore.CYAN + "  " + " " * indent + line)

    print()
    # 打印中间分隔线
    print(Style.BRIGHT + Fore.YELLOW + "-" * (max_length + 4))
    print()

    # 信息区域，使用不同颜色区分
    info = [
        (Fore.GREEN, "Author: bifish"),
        (Fore.YELLOW, "Name: IPHunter v1.0"),
        (Fore.MAGENTA, "Github: https://github.com/Bifish0")
    ]

    # 打印信息，保持居中对齐
    for color, text in info:
        # 计算信息的缩进
        info_indent = (max_length - len(text)) // 2
        print(Style.BRIGHT + color + "  " + " " * info_indent + text)

    print()
    # 打印底部边框
    print(Style.BRIGHT + Fore.YELLOW + "=" * (max_length + 4))
    print()


class IPDetector:
    """IP检测工具类，支持IP:端口格式处理"""

    # 页面元素定位（提供多种定位方式作为备选）
    INPUT_LOCATORS = [
        (By.ID, 'ipInput'),
        (By.XPATH, '//*[@id="ipInput"]'),
        (By.CSS_SELECTOR, 'input[name="ip"]'),
        (By.CSS_SELECTOR, 'input[type="text"][placeholder*="IP"]')
    ]

    SEARCH_LOCATORS = [
        (By.ID, 'searchBtn'),
        (By.XPATH, '//*[@id="searchBtn"]'),
        (By.CSS_SELECTOR, 'button[type="button"]'),
        (By.XPATH, '//button[contains(text(), "查询")]')
    ]

    RESULT_LOCATORS = [
        (By.XPATH, '/html/body/div[1]/section[1]/div/table/tbody/tr[7]/td[2]'),
        (By.XPATH, '//tr[contains(td/text(), "使用类型")]/td[2]'),
        (By.CSS_SELECTOR, 'table tbody tr:nth-child(7) td:nth-child(2)')
    ]

    COUNTRY_LOCATORS = [
        (By.XPATH, '/html/body/div[1]/section[1]/div/table/tbody/tr[2]/td[2]'),
        (By.XPATH, '//tr[contains(td/text(), "国家/地区")]/td[2]'),
        (By.CSS_SELECTOR, 'table tbody tr:nth-child(2) td:nth-child(2)')
    ]

    # 检测目标网站
    TARGET_URL = "https://ipip.la/"

    def __init__(self):
        self.driver = None
        self.total_count = 0
        self.success_count = 0
        self.error_count = 0
        self.invalid_ip_count = 0
        self.start_time = 0
        self.workbook = None
        self.worksheet = None

    def print_title(self):
        """打印美化的标题"""
        title = "IP 地址检测工具 v1.0"
        line = "=" * (len(title) + 9)
        print(f"\n{Fore.CYAN}{line}")
        print(f" {Fore.YELLOW}{Style.BRIGHT}{title} ")
        print(f"{Fore.CYAN}{line}")
        print(f"{Fore.GREEN}检测网站: {self.TARGET_URL}")
        print(f"{Fore.GREEN}开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

    def init_browser(self):
        """初始化浏览器配置，使用无头模式并隐藏自动化特征"""
        # 检查驱动是否存在
        system = platform.system()
        if system == "Windows":
            chrome_driver_name = 'chromedriver.exe'
        elif system == "Darwin":  # macOS
            chrome_driver_name = 'chromedriver'
        else:  # Linux
            chrome_driver_name = 'chromedriver'

        chrome_driver_path = os.path.join(os.getcwd(), chrome_driver_name)
        if not os.path.exists(chrome_driver_path):
            print(f"{Fore.RED}错误: 未找到 {chrome_driver_name} 文件")
            print(f"{Fore.YELLOW}提示: 请确保Chrome驱动程序与当前Chrome版本匹配，并放在程序同一目录下")
            print(f"{Fore.YELLOW}下载地址: https://sites.google.com/chromium.org/driver/")
            return False

        # 浏览器配置 - 隐藏自动化特征并启用无头模式
        chrome_options = Options()

        # 启用无头模式
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--disable-gpu")

        # 基础优化配置
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disk-cache-size=104857600")
        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument("--disable-background-networking")
        chrome_options.add_argument("--disable-features=TranslateUI")

        # 隐藏自动化控制特征
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option("useAutomationExtension", False)

        # 页面加载策略
        chrome_options.page_load_strategy = 'eager'

        # 初始化WebDriver
        try:
            service = Service(chrome_driver_path)
            service.log_path = os.devnull  # 禁用日志

            self.driver = webdriver.Chrome(service=service, options=chrome_options)

            # 进一步隐藏自动化特征
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    })
                """
            })

            self.driver.set_page_load_timeout(15)
            self.driver.set_script_timeout(10)
            return True
        except WebDriverException as e:
            print(f"{Fore.RED}浏览器初始化失败: {str(e)}")
            print(f"{Fore.YELLOW}可能原因: Chrome驱动与浏览器版本不匹配")
            return False
        except Exception as e:
            print(f"{Fore.RED}浏览器初始化失败: {str(e)}")
            return False

    def is_valid_ip(self, ip):
        """验证IP地址格式是否有效"""
        # 移除可能的端口号
        if ':' in ip:
            ip = ip.split(':')[0]

        # 校验IPv4格式
        ipv4_pattern = r'^((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
        if re.match(ipv4_pattern, ip):
            return True

        # 可扩展添加IPv6校验
        return False

    def load_ip_list(self):
        """加载并验证IP列表从ip.txt文件，支持ip:端口格式"""
        try:
            with open('ip.txt', 'r', encoding='utf-8') as f:
                raw_list = [line.strip() for line in f if line.strip()]

            if not raw_list:
                print(f"{Fore.YELLOW}警告: ip.txt文件为空，没有需要检测的IP")
                return []

            # 验证IP格式，区分完整格式(含端口)和纯IP
            valid_ips = []  # 存储元组 (原始格式, 纯IP)
            invalid_ips = []
            for item in raw_list:
                # 提取IP部分
                if ':' in item:
                    ip_part = item.split(':')[0]
                else:
                    ip_part = item

                if self.is_valid_ip(ip_part):
                    valid_ips.append((item, ip_part))  # 保留原始格式和纯IP
                else:
                    invalid_ips.append(item)

            self.invalid_ip_count = len(invalid_ips)
            self.total_count = len(valid_ips)

            print(f"{Fore.GREEN}成功加载 {self.total_count} 个有效IP地址")
            if self.invalid_ip_count > 0:
                print(f"{Fore.YELLOW}检测到 {self.invalid_ip_count} 个无效IP格式，已自动过滤：")
                for ip in invalid_ips[:5]:  # 只显示前5个
                    print(f"  - {ip}")
                if len(invalid_ips) > 5:
                    print(f"  - 以及 {len(invalid_ips) - 5} 个更多...")
            print(f"\n准备开始检测...\n")
            return valid_ips
        except FileNotFoundError:
            print(f"{Fore.RED}错误: 未找到ip.txt文件")
            print(f"{Fore.YELLOW}提示: 请在程序同一目录下创建ip.txt，并在其中每行填写一个IP地址(可带端口)")
            return []
        except Exception as e:
            print(f"{Fore.RED}读取IP列表失败: {str(e)}")
            return []

    def find_element_with_fallback(self, locators, timeout=10):
        """使用多个定位器查找元素，提供备选方案"""
        for by, value in locators:
            try:
                return WebDriverWait(self.driver, timeout, poll_frequency=0.5).until(
                    EC.presence_of_element_located((by, value))
                )
            except (TimeoutException, NoSuchElementException):
                continue
        return None

    def find_clickable_element_with_fallback(self, locators, timeout=10):
        """使用多个定位器查找可点击元素，提供备选方案"""
        for by, value in locators:
            try:
                return WebDriverWait(self.driver, timeout, poll_frequency=0.5).until(
                    EC.element_to_be_clickable((by, value))
                )
            except (TimeoutException, NoSuchElementException):
                continue
        return None

    def safe_element_operation(self, operation, max_retries=3):
        """增强元素操作稳定性，增加重试次数"""
        for attempt in range(max_retries):
            try:
                return operation()
            except StaleElementReferenceException:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
                return False, "元素已失效"
            except TimeoutException:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                return False, "操作超时"
            except Exception as e:
                return False, f"操作失败: {str(e)}"
        return False, "达到最大重试次数"

    def detect_ip(self, ip):
        """优化检测逻辑，提高单个IP检测成功率"""

        # 输入IP地址
        def input_operation():
            input_box = self.find_clickable_element_with_fallback(self.INPUT_LOCATORS, 5)
            if not input_box:
                return False, "未找到输入框"

            input_box.clear()
            input_box.send_keys(ip)
            return True, None

        success, error = self.safe_element_operation(input_operation)
        if not success:
            return False, error, ""

        # 点击搜索按钮
        def click_operation():
            search_btn = self.find_clickable_element_with_fallback(self.SEARCH_LOCATORS, 5)
            if not search_btn:
                return False, "未找到搜索按钮"

            search_btn.click()
            time.sleep(0.5)  # 等待请求发出
            return True, None

        success, error = self.safe_element_operation(click_operation)
        if not success:
            return False, error, ""

        # 获取国家/地区信息
        def get_country_operation():
            country_element = self.find_element_with_fallback(self.COUNTRY_LOCATORS, 8)
            if not country_element:
                return False, "未找到国家/地区信息"
            return True, country_element.text.strip()

        country_success, country = self.safe_element_operation(get_country_operation)
        if not country_success:
            return False, "获取国家/地区失败", ""

        # 获取IP类型结果
        def get_result_operation():
            result_element = self.find_element_with_fallback(self.RESULT_LOCATORS, 8)
            if not result_element:
                return False, "未找到IP类型信息"

            # 获取原始文本并处理
            raw_text = result_element.text.strip()
            # 移除任何可能存在的"正在同步数据"文本
            cleaned_text = raw_text.replace("正在同步数据", "").strip()
            return True, cleaned_text

        result_success, ip_type = self.safe_element_operation(get_result_operation)
        if not result_success:
            return False, "获取IP类型失败", country

        return True, ip_type, country

    def print_result(self, original_ip, country, ip_type, index):
        """打印检测结果，按要求设置特定IP类型文字颜色"""
        # 定义需要高亮的关键词及对应颜色
        highlighted_type = ip_type
        # 绿色文字
        if "原生IP" in highlighted_type:
            highlighted_type = highlighted_type.replace("原生IP", f"{Fore.LIGHTGREEN_EX}原生IP{Style.RESET_ALL}")
        if "家庭带宽IP" in highlighted_type:
            highlighted_type = highlighted_type.replace("家庭带宽IP",
                                                        f"{Fore.LIGHTGREEN_EX}家庭带宽IP{Style.RESET_ALL}")
        # 橙色文字
        if "IDC机房IP" in highlighted_type:
            highlighted_type = highlighted_type.replace("IDC机房IP", f"{Fore.LIGHTYELLOW_EX}IDC机房IP{Style.RESET_ALL}")
        if "广播IP" in highlighted_type:
            highlighted_type = highlighted_type.replace("广播IP", f"{Fore.LIGHTYELLOW_EX}广播IP{Style.RESET_ALL}")
        # 灰色文字
        if "未知类型" in highlighted_type:
            highlighted_type = highlighted_type.replace("未知类型", f"{Fore.LIGHTBLACK_EX}未知类型{Style.RESET_ALL}")

        print(
            f"{Fore.WHITE}{Style.BRIGHT}{original_ip:<20}  {Fore.RED}国家/地区: {Style.BRIGHT}{country:<15}  {Fore.BLUE}IP类型: {Style.BRIGHT}{highlighted_type}")

        # 打印进度信息
        elapsed = time.time() - self.start_time
        avg_time = elapsed / index
        progress = f"[{index}/{self.total_count}] 总耗时: {elapsed:.2f}s  平均: {avg_time:.2f}s/个"
        print(f"{Fore.YELLOW}{progress}\n")

    def print_error(self, original_ip, error_msg, index):
        """打印错误信息"""
        print(f"{Fore.WHITE}{Style.BRIGHT}{original_ip:<20}  {Fore.RED}{error_msg}")

        # 打印进度信息
        elapsed = time.time() - self.start_time
        avg_time = elapsed / index
        progress = f"[{index}/{self.total_count}] 总耗时: {elapsed:.2f}s  平均: {avg_time:.2f}s/个"
        print(f"{Fore.YELLOW}{progress}\n")

    def init_excel(self):
        """初始化Excel工作簿并美化"""
        self.workbook = Workbook(write_only=False)  # 可读写模式以支持样式
        self.worksheet = self.workbook.active
        self.worksheet.title = "IP检测结果"

        # 设置表头 - 修改为host和IP地址两列
        self.worksheet.append(["序号", "host", "IP地址", "国家/地区", "IP类型"])

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 设置表头样式和列宽
        for col in range(1, 6):  # 5列数据
            cell = self.worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # 调整列宽
        self.worksheet.column_dimensions['A'].width = 8
        self.worksheet.column_dimensions['B'].width = 30  # host列宽度
        self.worksheet.column_dimensions['C'].width = 20  # IP地址列宽度
        self.worksheet.column_dimensions['D'].width = 20  # 国家/地区列宽度
        self.worksheet.column_dimensions['E'].width = 30  # IP类型列宽度

        # 保存表头样式设置
        self.header_font = header_font
        self.cell_alignment = cell_alignment
        self.border = border

    def save_result(self, original_ip, country, ip_type):
        """保存结果到Excel，应用单元格样式"""
        try:
            row = self.success_count + self.error_count + 1  # +1因为表头占了一行

            # 提取纯IP部分
            if ':' in original_ip:
                pure_ip = original_ip.split(':')[0]
                # 转换为http://ip:port格式
                host = f"http://{original_ip}"
            else:
                pure_ip = original_ip
                host = original_ip  # 没有端口的情况保持原样

            # 添加数据行
            self.worksheet.append([row, host, pure_ip, country, ip_type])

            # 应用样式到新添加的行
            for col in range(1, 6):
                cell = self.worksheet.cell(row=row + 1, column=col)  # +1因为表头占了一行
                cell.alignment = self.cell_alignment
                cell.border = self.border

                # 为IP类型列添加特殊样式
                if col == 5:
                    if "原生IP" in ip_type or "家庭带宽IP" in ip_type:
                        cell.font = Font(color="008000")  # 绿色文字
                    elif "IDC机房IP" in ip_type or "广播IP" in ip_type:
                        cell.font = Font(color="FFA500")  # 橙色文字
                    elif "未知类型" in ip_type:
                        cell.font = Font(color="808080")  # 灰色文字
        except Exception as e:
            print(f"{Fore.RED}写入结果文件失败: {str(e)}")

    def save_excel(self):
        """保存Excel文件"""
        try:
            filename = f"ip_result_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.workbook.save(filename)
            return filename
        except Exception as e:
            print(f"{Fore.RED}保存Excel文件失败: {str(e)}")
            return None

    def test_website_connectivity(self):
        """测试目标网站的基本网络连接性"""
        try:
            domain = self.TARGET_URL.split("//")[-1].split("/")[0]
            ip_address = socket.gethostbyname(domain)
            print(f"{Fore.YELLOW}域名解析成功: {domain} -> {ip_address}")

            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(5)
                result = s.connect_ex((ip_address, 443))
                if result == 0:
                    print(f"{Fore.YELLOW}TCP连接成功，端口443可访问")
                    return True
                else:
                    print(f"{Fore.YELLOW}TCP连接失败，端口443不可访问")
                    return False
        except socket.gaierror:
            print(f"{Fore.YELLOW}域名解析失败")
            return False
        except socket.timeout:
            print(f"{Fore.YELLOW}连接超时")
            return False
        except Exception as e:
            print(f"{Fore.YELLOW}网络测试失败: {str(e)}")
            return False

    def run(self):
        """运行检测主流程"""
        # 首先打印banner
        print_banner()
        # 然后打印标题信息
        self.print_title()

        # 初始化浏览器
        if not self.init_browser():
            return

        # 加载IP列表，获取(原始格式, 纯IP)元组列表
        ip_list = self.load_ip_list()
        if not ip_list:
            self.driver.quit()
            return

        # 初始化Excel文件
        self.init_excel()

        # 打开检测网站
        try:
            print(f"{Fore.CYAN}正在连接 {self.TARGET_URL}...")
            self.driver.get(self.TARGET_URL)

            # 等待关键元素加载
            input_box = self.find_clickable_element_with_fallback(self.INPUT_LOCATORS, 15)
            if not input_box:
                raise Exception("未找到输入框元素，网站可能已更新")

            print(f"{Fore.GREEN}成功连接网站，开始检测\n")

        except TimeoutException:
            print(f"{Fore.RED}错误: 连接网站超时")
            self.test_website_connectivity()
            self.driver.quit()
            return
        except WebDriverException as e:
            print(f"{Fore.RED}错误: 连接网站失败: {str(e)}")
            self.test_website_connectivity()
            self.driver.quit()
            return
        except Exception as e:
            print(f"{Fore.RED}错误: 连接网站时发生错误: {str(e)}")
            self.driver.quit()
            return

        # 开始检测计时
        self.start_time = time.time()

        # 逐个检测IP，增加重试机制
        for index, (original_ip, pure_ip) in enumerate(ip_list, 1):
            success = False
            result = ""
            country = ""

            # 最多重试3次
            for attempt in range(3):
                try:
                    # 使用纯IP进行检测
                    current_success, current_result, current_country = self.detect_ip(pure_ip)
                    if current_success:
                        success = True
                        result = current_result
                        country = current_country
                        break

                    # 遇到严重错误需要刷新页面
                    if "未找到输入框" in current_result or "未找到搜索按钮" in current_result:
                        print(f"{Fore.YELLOW}页面可能已变化，尝试刷新...")
                        self.driver.refresh()
                        time.sleep(2)
                        # 重新等待页面加载
                        self.find_clickable_element_with_fallback(self.INPUT_LOCATORS, 10)
                except Exception as e:
                    print(f"{Fore.YELLOW}第{attempt + 1}次尝试失败: {str(e)}")
                    time.sleep(1)

            if success:
                self.success_count += 1
                # 打印和保存时使用原始IP:端口格式
                self.print_result(original_ip, country, result, index)
                self.save_result(original_ip, country, result)
            else:
                self.error_count += 1
                self.print_error(original_ip, result if result else "检测失败", index)
                self.save_result(original_ip, country, result)

            # 智能调整请求间隔，避免触发频率限制
            elapsed = time.time() - self.start_time
            avg_time = elapsed / index

            # 根据平均耗时动态调整等待时间
            if avg_time < 1.5:
                wait_time = max(0.5, 1.5 - avg_time)
                time.sleep(wait_time)
            elif avg_time > 3:
                # 速度过慢则不额外等待
                pass
            else:
                time.sleep(0.3)

        # 清理资源
        excel_file = self.save_excel()
        self.driver.quit()

        # 打印总结信息
        total_time = time.time() - self.start_time
        print(f"{Fore.BLUE}" + "=" * 60)
        print(f"{Fore.YELLOW}检测完成! 总计: {self.total_count} 个")
        print(f"{Fore.YELLOW}成功: {self.success_count} 个 | 失败: {self.error_count} 个")
        print(f"{Fore.YELLOW}无效IP: {self.invalid_ip_count} 个")
        print(f"{Fore.YELLOW}总耗时: {total_time:.2f}s | 平均速度: {total_time / self.total_count:.2f}s/个")
        if excel_file:
            print(f"{Fore.YELLOW}结果文件: {os.path.abspath(excel_file)}")
        print(f"{Fore.BLUE}" + "=" * 60)


if __name__ == "__main__":
    detector = IPDetector()
    detector.run()
